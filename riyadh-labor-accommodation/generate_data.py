# -*- coding: utf-8 -*-
"""
مولّد البيانات التجريبية — لوحة معلومات سكن العمالة، أمانة منطقة الرياض
Generates labor_accommodation_data.xlsx (19 sheets) + data.json (embedded into index.html).

Seed totals (hard requirements):
  TOTAL_DEMAND (laborers)      = 1,420,000
  TOTAL_BEDS   (licensed beds) = 612,400   -> gap 807,600, coverage 43.1٪
  AVG_OCCUPANCY                = 91.7٪     -> occupied beds 561,571
Every breakdown reconciles exactly; assertions at the bottom enforce it.
"""
import json, math, random, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)
HERE = os.path.dirname(os.path.abspath(__file__))

TOTAL_DEMAND = 1_420_000
TOTAL_BEDS   = 612_400
TOTAL_OCC    = 561_571            # 91.7٪ of beds
TODAY        = "2026-07-08"

SECTORS = ["القطاع الشمالي", "القطاع الشرقي", "القطاع الأوسط", "القطاع الغربي", "القطاع الجنوبي"]
SEC_EN   = {"القطاع الشمالي":"NORTH","القطاع الشرقي":"EAST","القطاع الأوسط":"CENTER","القطاع الغربي":"WEST","القطاع الجنوبي":"SOUTH"}

AR_MONTHS = ["يناير","فبراير","مارس","أبريل","مايو","يونيو","يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]
def month_seq(start_y, start_m, n):
    out=[]; y,m=start_y,start_m
    for _ in range(n):
        out.append((y,m,f"{y}-{m:02d}",f"{AR_MONTHS[m-1]} {y}"))
        m+=1
        if m==13: m=1; y+=1
    return out
MONTHS24 = month_seq(2024,7,24)          # يوليو 2024 → يونيو 2026
FCAST12  = month_seq(2026,7,12)          # يوليو 2026 → يونيو 2027

def largest_remainder(total, weights):
    """Split integer total by weights, exact sum."""
    s=sum(weights); raw=[total*w/s for w in weights]
    base=[int(x) for x in raw]; rem=total-sum(base)
    order=sorted(range(len(raw)), key=lambda i: raw[i]-base[i], reverse=True)
    for i in range(rem): base[order[i%len(base)]]+=1
    return base

# ----------------------------------------------------------------------------
# 1-6  DEMAND BREAKDOWNS (each sums to exactly 1,420,000)
# ----------------------------------------------------------------------------
BLUE_TOTAL, WHITE_TOTAL = 1_164_400, 255_600   # 82٪ / 18٪

demand_occupation = [  # SSCO major groups
    ("المهن الأولية (العمالة العادية)",              "ياقات زرقاء", 498_000),
    ("الحرفيون وذوو المهن المرتبطة بهم",             "ياقات زرقاء", 312_000),
    ("مشغلو الآلات والمعدات وسائقوها",               "ياقات زرقاء", 176_400),
    ("عمال الخدمات والمبيعات",                        "ياقات زرقاء", 118_000),
    ("العاملون المهرة في الزراعة",                    "ياقات زرقاء",  60_000),
    ("الفنيون ومساعدو الاختصاصيين",                  "ياقات بيضاء",  96_600),
    ("الاختصاصيون",                                   "ياقات بيضاء",  74_000),
    ("موظفو الدعم الإداري (الكتبة)",                  "ياقات بيضاء",  55_000),
    ("المديرون",                                      "ياقات بيضاء",  30_000),
]
def with_collar_split(rows, blue_shares):
    """rows: [(name,total)], blue_shares: per-row blue fraction. Returns rows with exact blue/white, reconciled to city blue/white totals."""
    blues=[round(t*b) for (_,t),b in zip(rows,blue_shares)]
    diff=BLUE_TOTAL-sum(blues); blues[0]+=diff
    out=[]
    for (name,total),blue in zip(rows,blues):
        out.append({"name":name,"total":total,"blue":blue,"white":total-blue})
    return out

demand_econ = with_collar_split([
    ("التشييد والبناء",             505_000),
    ("التجارة والتجزئة",            217_000),
    ("الصناعة",                     198_000),
    ("الخدمات العامة",              165_000),
    ("الضيافة والإعاشة",            152_000),
    ("النقل والخدمات اللوجستية",    128_000),
    ("الزراعة",                      55_000),
], [0.93, 0.62, 0.85, 0.70, 0.78, 0.86, 0.95])

demand_sme = with_collar_split([
    ("كبيرة",          642_000),
    ("متوسطة",         388_000),
    ("صغيرة",          262_000),
    ("متناهية الصغر",  128_000),
], [0.84, 0.82, 0.79, 0.76])

demand_nat = with_collar_split([
    ("الهند",      362_000), ("بنغلاديش", 289_000), ("باكستان", 224_000),
    ("مصر",        141_000), ("اليمن",     98_000), ("الفلبين",  82_000),
    ("نيبال",       74_000), ("السودان",   55_000), ("إثيوبيا",  39_000),
    ("أخرى",        56_000),
], [0.82, 0.90, 0.85, 0.70, 0.72, 0.68, 0.92, 0.78, 0.88, 0.75])

demand_age = with_collar_split([
    ("18–24", 213_000), ("25–34", 561_000), ("35–44", 398_000),
    ("45–54", 186_000), ("55+",    62_000),
], [0.85, 0.83, 0.81, 0.79, 0.74])

SEC_DEMAND = {"القطاع الشمالي":258_000,"القطاع الشرقي":331_000,"القطاع الأوسط":246_000,"القطاع الغربي":217_000,"القطاع الجنوبي":368_000}
demand_sector = with_collar_split([(s,SEC_DEMAND[s]) for s in SECTORS], [0.79,0.83,0.76,0.82,0.87])

# ----------------------------------------------------------------------------
# 7  FACILITIES — 140 rows, beds sum = 612,400 exactly, occupied = 561,571
# ----------------------------------------------------------------------------
SEC_BEDS = {"القطاع الشمالي":128_000,"القطاع الشرقي":172_400,"القطاع الأوسط":96_000,"القطاع الغربي":88_000,"القطاع الجنوبي":128_000}
SEC_OCC  = {"القطاع الشمالي":113_280,"القطاع الشرقي":164_470,"القطاع الأوسط":86_592,"القطاع الغربي":77_933,"القطاع الجنوبي":119_296}
assert sum(SEC_BEDS.values())==TOTAL_BEDS and sum(SEC_OCC.values())==TOTAL_OCC

# (sector) -> [(type, count, beds_target)]
FAC_PLAN = {
 "القطاع الشمالي":[("مجمع سكني",8,112_000),("مبنى سكني",14,12_000),("كبائن متنقلة",8,4_000)],
 "القطاع الشرقي":[("مجمع سكني",10,156_000),("مبنى سكني",16,12_400),("كبائن متنقلة",8,4_000)],
 "القطاع الأوسط":[("مجمع سكني",6,84_600),("مبنى سكني",13,9_000),("كبائن متنقلة",7,2_400)],
 "القطاع الغربي":[("مجمع سكني",6,78_400),("مبنى سكني",11,8_000),("كبائن متنقلة",5,1_600)],
 "القطاع الجنوبي":[("مجمع سكني",8,115_600),("مبنى سكني",14,10_200),("كبائن متنقلة",6,2_200)],
}
RANGES = {"مجمع سكني":(3000,22000),"مبنى سكني":(200,1200),"كبائن متنقلة":(80,600)}

COMP_NAMES=["النخيل","الواحة","الرمال الذهبية","السلام","المروج","اليمامة","الديرة","الوسام","الفيصلية","الأصالة",
 "البستان","الروابي","الصفوة","الرائد","المجد","الازدهار","النهضة","التعمير","الإنجاز","الأفق",
 "المستقبل","البنيان","الأمان","النخبة","الديار","السدرة","الغضا","الطلح","العارض","حنيفة",
 "الشعلة","التلال","المها","الجزيرة","الصحراء","الفرسان","العزم","الريادة"]
BLDG_NAMES=["الياقوت","المرجان","اللؤلؤ","الزمرد","الندى","الفجر","الشروق","الغدير","الينبوع","الرابية",
 "الصفا","المروة","التوت","الأرجوان","الزيتون","السرو","الغيم","المطر","البرق","السحاب",
 "الوسيل","الجوهرة","الماسة","العقيق","الفيروز","البلور","الصخور","الحصى","الرمث","العرفج",
 "السمر","الأثل","الطلحة","السلم","القرط","الخزامى","الأقحوان","النرجس","الريحان","الزنبق",
 "البنفسج","الياسمين","الورد","الفل","الكادي","العنبر","المسك","العود","الصندل","الزعفران",
 "النعناع","الحبق","الآس","الروض","المرج","السهل","الوادي","الهضبة","الرمل","الكثيب",
 "البيداء","الفلاة","النفود","الدهناء","الصمان","العرمة","التنهات","البرة"]
CABIN_NAMES=["مشروع المطار","وادي حنيفة","طريق الخرج","المنطقة الصناعية الثانية","مشروع المترو الجنوبي",
 "ضاحية عرقة","مخطط الرمال","طريق ديراب","مشروع تصريف السيول","المدينة الصناعية",
 "مخرج 18","مخرج 25","طريق الثمامة","مشروع الملك سلمان","حي السلي الصناعي",
 "مستودعات السلي","طريق المزاحمية","مشروع النقل العام","محطة المعالجة","مخطط بدر",
 "الحاير","العزيزية الصناعية","طريق الدمام السريع","مجمع الورش الشرقي","مخرج 30",
 "طريق الجنادرية","مشروع الأنفاق","حي المصفاة","المرقب الجنوبي","مشروع الإسكان الشرقي",
 "طريق الحائر","مخطط الشفا","مشروع جسر النسيم","المنطقة اللوجستية"]

# X/Y sampling boxes per sector (inside the SVG polygons defined in GEO below)
BOX = {"القطاع الشمالي":(19,44,15,42),"القطاع الشرقي":(56,90,10,42),"القطاع الأوسط":(43,55,48,62),
       "القطاع الغربي":(10,34,54,72),"القطاع الجنوبي":(47,79,57,89)}

DISTRICTS = {
 "القطاع الشمالي":["الملقا","الياسمين","النرجس","الصحافة","حطين"],
 "القطاع الشرقي":["النسيم","الروضة","الريان","إشبيلية","القادسية"],
 "القطاع الأوسط":["الملز","العليا","المربع","الديرة","الوزارات"],
 "القطاع الغربي":["عرقة","ظهرة لبن","السويدي الغربي","العوالي","نمار"],
 "القطاع الجنوبي":["السلي","المنصورية","العزيزية","الدار البيضاء","الشفا"],
}

def spread_int(total, n, lo, hi):
    """n integers in [lo,hi] summing to total."""
    mean=total/n
    vals=[max(lo,min(hi,round(mean*random.uniform(0.55,1.45)))) for _ in range(n)]
    diff=total-sum(vals)
    guard=0
    while diff!=0 and guard<10000:
        i=random.randrange(n); step=max(1,abs(diff)//n)
        if diff>0 and vals[i]+step<=hi: vals[i]+=step; diff-=step
        elif diff<0 and vals[i]-step>=lo: vals[i]-=step; diff+=step
        guard+=1
    assert diff==0
    return vals

facilities=[]; fid=1000
comp_i=bld_i=cab_i=0
for sec in SECTORS:
    for ftype,count,target in FAC_PLAN[sec]:
        lo,hi=RANGES[ftype]
        beds=spread_int(target,count,lo,hi)
        # occupancy draws — الشرقي pressured (many >95٪), others 82–96٪
        for b in beds:
            if ftype=="مجمع سكني":
                name=f"مجمع {COMP_NAMES[comp_i%len(COMP_NAMES)]} لسكن العمالة"; comp_i+=1
            elif ftype=="مبنى سكني":
                name=f"مبنى {BLDG_NAMES[bld_i%len(BLDG_NAMES)]} السكني"; bld_i+=1
            else:
                name=f"كبائن {CABIN_NAMES[cab_i%len(CABIN_NAMES)]}"; cab_i+=1
            x0,x1,y0,y1=BOX[sec]
            facilities.append({
                "id":f"F-{fid}","name":name,"type":ftype,"sector":sec,
                "district":random.choice(DISTRICTS[sec]),
                "x":round(random.uniform(x0,x1),1),"y":round(random.uniform(y0,y1),1),
                "beds":b,
            }); fid+=1

# occupied beds per sector reconciled exactly
for sec in SECTORS:
    fs=[f for f in facilities if f["sector"]==sec]
    target=SEC_OCC[sec]; base_rate=target/SEC_BEDS[sec]
    occ=[]
    for f in fs:
        r=base_rate+random.uniform(-0.055,0.055)
        if sec=="القطاع الشرقي" and f["type"]=="مجمع سكني" and random.random()<0.6:
            r=random.uniform(0.955,0.985)
        r=max(0.78,min(0.985,r))
        occ.append(min(f["beds"],round(f["beds"]*r)))
    diff=target-sum(occ); guard=0
    while diff!=0 and guard<20000:
        i=random.randrange(len(fs))
        cap=fs[i]["beds"]; step=max(1,abs(diff)//len(fs))
        if diff>0 and occ[i]+step<=cap and (occ[i]+step)/cap<=0.99: occ[i]+=step; diff-=step
        elif diff<0 and (occ[i]-step)/cap>=0.75: occ[i]-=step; diff+=step
        guard+=1
    assert diff==0, sec
    for f,o in zip(fs,occ):
        f["occupied"]=o; f["occ_rate"]=round(o/f["beds"]*100,1)

# license status/dates + compliance + risk + last inspection
lic_year_pool=[2023,2024,2024,2025,2025,2025,2026]
expiring=random.sample(range(len(facilities)),12)
for i,f in enumerate(facilities):
    y=random.choice(lic_year_pool); m=random.randint(1,12); d=random.randint(1,28)
    if y==2026: m=random.randint(1,6)
    f["lic_status"]="منتهٍ قريباً" if i in expiring else "ساري"
    f["lic_date"]=f"{y}-{m:02d}-{d:02d}"
    comp=random.randint(62,98)
    if f["sector"]=="القطاع الجنوبي": comp=min(comp,random.randint(60,88))
    if i in expiring: comp=random.randint(62,74)
    f["compliance"]=comp
    f["risk"]="مرتفع" if (comp<70 or f["occ_rate"]>96) else ("متوسط" if (comp<82 or f["occ_rate"]>92) else "منخفض")
    im=random.randint(4,6); idd=random.randint(1,28)
    f["last_visit"]=f"2026-{im:02d}-{min(idd, 28):02d}"

# ----------------------------------------------------------------------------
# 8  LICENSES_MONTHLY — 24 months × 5 sectors × 3 types = 360 rows
# ----------------------------------------------------------------------------
SEC_LIC_SHARE={"القطاع الشرقي":0.26,"القطاع الجنوبي":0.24,"القطاع الشمالي":0.22,"القطاع الأوسط":0.15,"القطاع الغربي":0.13}
TYPE_SHARE={"مجمع سكني":0.30,"مبنى سكني":0.48,"كبائن متنقلة":0.22}
TYPES=list(TYPE_SHARE.keys())

licenses=[]
for mi,(y,m,key,label) in enumerate(MONTHS24):
    growth=(1.18)**(mi/12.0)
    season=1+0.06*math.sin(2*math.pi*((m-1)/12.0)+0.8)
    c_city=round(118*growth*season*random.uniform(0.95,1.05))
    o_city=round(158*growth*season*random.uniform(0.95,1.05))
    req_city=round(465*(1.12)**(mi/12.0)*season*random.uniform(0.96,1.04))
    close_rate=0.82+0.055*(mi/23.0)+random.uniform(-0.012,0.012)
    closed_city=round(req_city*close_rate)
    c_secs=largest_remainder(c_city,[SEC_LIC_SHARE[s] for s in SECTORS])
    o_secs=largest_remainder(o_city,[SEC_LIC_SHARE[s] for s in SECTORS])
    r_secs=largest_remainder(req_city,[SEC_LIC_SHARE[s] for s in SECTORS])
    cl_secs=largest_remainder(closed_city,[SEC_LIC_SHARE[s] for s in SECTORS])
    for si,sec in enumerate(SECTORS):
        d_off={"القطاع الجنوبي":3.5,"القطاع الشرقي":1.5,"القطاع الأوسط":0,"القطاع الغربي":-1,"القطاع الشمالي":-2}[sec]
        days_c=round(58-20*(mi/23.0)+d_off+random.uniform(-2.5,2.5),1)
        days_o=round(34-13*(mi/23.0)+d_off*0.6+random.uniform(-1.8,1.8),1)
        cts=largest_remainder(c_secs[si],[TYPE_SHARE[t] for t in TYPES])
        ots=largest_remainder(o_secs[si],[TYPE_SHARE[t] for t in TYPES])
        rts=largest_remainder(r_secs[si],[TYPE_SHARE[t] for t in TYPES])
        clts=largest_remainder(cl_secs[si],[TYPE_SHARE[t] for t in TYPES])
        for ti,t in enumerate(TYPES):
            licenses.append({"month":key,"month_ar":label,"sector":sec,"type":t,
                "constr":cts[ti],"ops":ots[ti],"req_in":rts[ti],"req_closed":min(clts[ti],rts[ti]),
                "days_constr":days_c,"days_ops":days_o})

# ----------------------------------------------------------------------------
# 9  INSPECTIONS_MONTHLY — 24 months × 5 sectors
# ----------------------------------------------------------------------------
SEC_INSP_SHARE={"القطاع الجنوبي":0.28,"القطاع الشرقي":0.24,"القطاع الشمالي":0.18,"القطاع الأوسط":0.17,"القطاع الغربي":0.13}
SEC_VIOL_MULT={"القطاع الجنوبي":1.42,"القطاع الشرقي":1.12,"القطاع الشمالي":0.72,"القطاع الأوسط":0.95,"القطاع الغربي":0.82}
inspections=[]
for mi,(y,m,key,label) in enumerate(MONTHS24):
    visits_city=round(1380*(1+0.25*(mi/23.0))*(1+0.04*math.sin(2*math.pi*((m-1)/12.0)))*random.uniform(0.97,1.03))
    viol_rate=0.212-0.054*(mi/23.0)
    v_secs=largest_remainder(visits_city,[SEC_INSP_SHARE[s] for s in SECTORS])
    for si,sec in enumerate(SECTORS):
        visits=v_secs[si]
        viol=round(visits*viol_rate*SEC_VIOL_MULT[sec]*random.uniform(0.9,1.1))
        fines=round(viol*random.uniform(7800,10500)/100)*100
        closures=max(0,round(viol*0.025+random.uniform(-0.8,1.6)))
        inspections.append({"month":key,"month_ar":label,"sector":sec,
            "visits":visits,"violations":viol,"fines":fines,"closures":closures})

TOT_VISITS=sum(r["visits"] for r in inspections)
TOT_VIOL=sum(r["violations"] for r in inspections)
TOT_FINES=sum(r["fines"] for r in inspections)
TOT_CLOSURES=sum(r["closures"] for r in inspections)

# ----------------------------------------------------------------------------
# 10  VIOLATION CATEGORIES (sum = TOT_VIOL) + SME split per category
# ----------------------------------------------------------------------------
CAT_DEFS=[("الاكتظاظ وتجاوز الطاقة الاستيعابية",0.31,13500),
          ("اشتراطات السلامة والحماية من الحريق",0.24,16000),
          ("النظافة والصحة العامة",0.17,6500),
          ("مزاولة النشاط بدون ترخيص",0.12,21000),
          ("مخالفات الكهرباء والسباكة",0.09,8500),
          ("التهوية والعزل غير المطابق",0.07,7000)]
cat_counts=largest_remainder(TOT_VIOL,[w for _,w,_ in CAT_DEFS])
SME_ORDER=["متناهية الصغر","صغيرة","متوسطة","كبيرة"]
SME_W=[0.14,0.33,0.34,0.19]
viol_cats=[]
for (name,_,avg_fine),cnt in zip(CAT_DEFS,cat_counts):
    sme=largest_remainder(cnt,SME_W if name!="مزاولة النشاط بدون ترخيص" else [0.30,0.40,0.22,0.08])
    viol_cats.append({"name":name,"count":cnt,
        "fines":round(cnt*avg_fine*random.uniform(0.93,1.07)/1000)*1000,
        "sme":{k:v for k,v in zip(SME_ORDER,sme)}})
# reconcile category fines to inspections total
fscale=TOT_FINES/sum(c["fines"] for c in viol_cats)
run=0
for c in viol_cats[:-1]:
    c["fines"]=round(c["fines"]*fscale/1000)*1000; run+=c["fines"]
viol_cats[-1]["fines"]=TOT_FINES-run

# by-sector split of each category (for the drawer)
for c in viol_cats:
    mult=[SEC_INSP_SHARE[s]*SEC_VIOL_MULT[s] for s in SECTORS]
    c["by_sector"]={s:v for s,v in zip(SECTORS,largest_remainder(c["count"],mult))}

# ----------------------------------------------------------------------------
# 11  INSPECTORS — 18, totals reconcile to 24-month visits/violations
# ----------------------------------------------------------------------------
INSPECTOR_NAMES=["م. عبدالله الشهراني","م. خالد العتيبي","م. فهد الدوسري","م. سعد القحطاني","م. محمد الحربي",
 "م. تركي المطيري","م. ناصر الغامدي","م. بندر الزهراني","م. سلطان السبيعي","م. عبدالعزيز العنزي",
 "م. راشد الشمري","م. يوسف البقمي","م. ماجد الرشيدي","م. وليد الجهني","م. عمر المالكي",
 "م. طلال العمري","م. حسام النفيعي","م. زياد الحقباني"]
insp_sector=[SECTORS[i%5] for i in range(18)]
w_visits=[random.uniform(0.8,1.25) for _ in range(18)]
insp_visits=largest_remainder(TOT_VISITS,w_visits)
insp_viol=largest_remainder(TOT_VIOL,[v*SEC_VIOL_MULT[s]*random.uniform(0.85,1.15) for v,s in zip(w_visits,insp_sector)])
inspectors=[]
for i in range(18):
    inspectors.append({"name":INSPECTOR_NAMES[i],"sector":insp_sector[i],
        "visits":insp_visits[i],"violations":insp_viol[i],
        "distance":round(insp_visits[i]*random.uniform(8.5,14.5))})

# ----------------------------------------------------------------------------
# 12  HOTSPOTS — 40 rows, skewed high in الجنوبي وأجزاء من الشرقي
# ----------------------------------------------------------------------------
HOT_PLAN=[("القطاع الجنوبي",12,(58,100)),("القطاع الشرقي",10,(35,92)),("القطاع الأوسط",7,(18,58)),
          ("القطاع الشمالي",6,(8,42)),("القطاع الغربي",5,(10,38))]
hotspots=[]
for sec,n,(lo,hi) in HOT_PLAN:
    x0,x1,y0,y1=BOX[sec]
    for _ in range(n):
        hotspots.append({"x":round(random.uniform(x0,x1),1),"y":round(random.uniform(y0,y1),1),
            "sector":sec,"density":random.randint(lo,hi)})

# ----------------------------------------------------------------------------
# 13  INITIATIVES — 28 across 7 pillars; 3 close within 90 days of TODAY
# ----------------------------------------------------------------------------
PILLARS=["الأنظمة والتشريعات","النموذج الحضري","الاستثمار والقطاع الخاص","الرقابة"]
OWNERS={"الأنظمة والتشريعات":"إدارة التنظيمات البلدية","النموذج الحضري":"الإدارة العامة للتخطيط العمراني",
 "الاستثمار والقطاع الخاص":"إدارة الاستثمار والشراكات","الرقابة":"الإدارة العامة للرقابة الميدانية"}
INIT_DEFS={
 "الأنظمة والتشريعات":[
  ("تحديث لائحة اشتراطات السكن الجماعي للأفراد","في المسار",78,"2025-03-01","2026-11-25",4_200_000,"اعتماد النسخة النهائية من اللجنة العليا","منخفض","رفع نسبة المنشآت المتوافقة إلى 85٪"),
  ("إطار تصنيف مقدمي خدمات الإسكان الجماعي","في المسار",55,"2025-06-15","2027-01-31",2_800_000,"إطلاق التصنيف التجريبي لـ 40 منشأة","متوسط","تحفيز رفع جودة المرافق عبر التصنيف المعلن"),
  ("توحيد الاشتراطات مع الجهات ذات العلاقة","متأخرة",42,"2025-01-10","2026-06-30",1_900_000,"توقيع مذكرة التفاهم مع وزارة الموارد البشرية والتنمية الاجتماعية","مرتفع","تقليص ازدواجية الاشتراطات بنسبة 60٪"),
  ("دليل الاشتراطات الفنية للكبائن المتنقلة","مكتملة",100,"2024-09-01","2025-11-30",950_000,"—","منخفض","تنظيم قطاع الكبائن المؤقتة بالكامل"),
  ("مصفوفة الصلاحيات والأدوار للجهات المشاركة","مكتملة",100,"2024-08-01","2025-09-30",700_000,"—","منخفض","وضوح المساءلة عبر 9 جهات"),
  ("سياسة تبادل البيانات مع الجهات الوطنية","متأخرة",45,"2025-04-01","2026-06-15",1_200_000,"توقيع اتفاقية التكامل مع المنصة الوطنية","مرتفع","تحديث تلقائي لبيانات العمالة شهرياً"),
  ("إطار قياس أثر المبادرات الاستراتيجية","في المسار",62,"2025-07-01","2027-01-15",900_000,"اعتماد منهجية القياس","متوسط","ربط الإنفاق بالأثر لكل مبادرة"),
 ],
 "النموذج الحضري":[
  ("مخطط المناطق النموذجية للسكن الجماعي (المرحلة الأولى)","في المسار",64,"2025-02-01","2027-03-31",38_000_000,"اعتماد المخطط التفصيلي للموقع الجنوبي","متوسط","إضافة 45,000 سرير مرخص جديد"),
  ("نقل التجمعات العشوائية من الأحياء السكنية","متأخرة",38,"2024-11-01","2026-12-31",22_500_000,"حصر التجمعات المستهدفة في السلي والمنصورية","مرتفع","إخلاء 120 موقعاً غير نظامي"),
  ("تأهيل البنية التحتية للمدينة العمالية بالمصفاة","في المسار",71,"2025-04-20","2026-12-10",17_300_000,"استكمال شبكات الصرف للمرحلة الثانية","متوسط","رفع الطاقة الاستيعابية 18,000 سرير"),
  ("دراسة مواءمة استعمالات الأراضي مع الطلب","مكتملة",100,"2024-06-01","2025-08-31",1_600_000,"—","منخفض","خريطة فجوة معتمدة لتوجيه التراخيص"),
  ("التوأم الرقمي لمواقع السكن الجماعي","في المسار",57,"2025-05-15","2027-05-31",9_800_000,"مسح ثلاثي الأبعاد لأول 200 منشأة","متوسط","خفض زيارات التحقق الميدانية 35٪"),
  ("مكتب إدارة بيانات السكن الجماعي","في المسار",74,"2025-03-10","2026-11-30",2_900_000,"إطلاق قاموس البيانات الموحد","منخفض","مصدر واحد موثوق لكل مؤشرات القطاع"),
  ("لوحات المؤشرات التنفيذية الموحدة","في المسار",84,"2025-06-01","2026-09-30",1_500_000,"ربط مصادر البيانات المتبقية","منخفض","قرارات أسرع بمعلومة موحدة"),
 ],
 "الاستثمار والقطاع الخاص":[
  ("طرح 6 فرص استثمارية لمجمعات السكن الجماعي","في المسار",59,"2025-05-01","2027-02-28",5_400_000,"إغلاق التأهيل المسبق للمطورين","متوسط","استقطاب استثمارات بـ 1.2 مليار ريال"),
  ("برنامج حوافز التراخيص للمطورين","في المسار",83,"2025-01-15","2026-11-15",3_100_000,"إطلاق حزمة الحوافز الثانية","منخفض","تسريع دخول 25,000 سرير للسوق"),
  ("شراكة تشغيل المجمعات الحكومية مع القطاع الخاص","متوقفة",22,"2025-03-01","2026-11-30",8_700_000,"إعادة طرح كراسة الشروط","مرتفع","رفع كفاءة التشغيل وخفض التكاليف 30٪"),
  ("منصة مطابقة العرض والطلب للمنشآت الصغيرة","في المسار",47,"2025-08-01","2027-04-30",2_300_000,"اكتمال النموذج الأولي للمنصة","متوسط","خدمة 8,000 منشأة صغيرة ومتناهية الصغر"),
  ("ملتقى الرياض السنوي للإسكان الجماعي","متوقفة",18,"2025-06-01","2026-11-15",3_200_000,"تحديد موعد وشركاء النسخة الأولى","متوسط","منصة سنوية للقطاعين العام والخاص"),
  ("حملة «سكن آمن» لتوعية أصحاب العمل","في المسار",66,"2025-04-01","2026-12-20",2_400_000,"إطلاق الموجة الثالثة من الحملة","منخفض","وصول 50,000 منشأة مستهدفة"),
  ("دليل المستأجر العامل بست لغات","مكتملة",100,"2025-02-01","2025-12-31",600_000,"—","منخفض","تمكين العمال من معرفة حقوقهم السكنية"),
 ],
 "الرقابة":[
  ("برنامج الرقابة الوقائية بالنمذجة التنبؤية","في المسار",68,"2025-02-15","2026-12-20",4_800_000,"تشغيل نموذج المخاطر على بيانات 6 أشهر","متوسط","رفع دقة استهداف الجولات إلى 75٪"),
  ("حملة الامتثال الشامل بالقطاع الجنوبي","في المسار",86,"2025-09-01","2026-08-15",3_500_000,"تغطية آخر 60 موقعاً عالي الخطورة","منخفض","خفض المخالفات الجسيمة 40٪ في الجنوب"),
  ("ربط أجهزة المفتشين بمنظومة بلدي الرقابية","مكتملة",100,"2024-10-01","2025-12-15",2_100_000,"—","منخفض","أتمتة 100٪ من محاضر الضبط"),
  ("إطار التفتيش المشترك مع الدفاع المدني","متأخرة",51,"2025-01-20","2026-11-30",1_400_000,"اعتماد بروتوكول الجولات المشتركة","مرتفع","توحيد ضبط اشتراطات السلامة"),
  ("منصة الترخيص الإلكتروني الموحد للسكن الجماعي","في المسار",91,"2024-12-01","2026-08-25",12_500_000,"إطلاق خدمة رخص البناء رقمياً","منخفض","إصدار 95٪ من الرخص إلكترونياً"),
  ("تطبيق البلاغات المجتمعية عن السكن العشوائي","مكتملة",100,"2025-01-01","2026-02-28",1_800_000,"—","منخفض","استقبال 1,500 بلاغ نوعي سنوياً"),
  ("برنامج سفراء الامتثال في المنشآت الكبرى","في المسار",49,"2025-09-15","2027-03-31",1_100_000,"تدريب الدفعة الثانية (120 سفيراً)","متوسط","قناة امتثال ذاتي داخل أكبر 500 منشأة"),
 ],
}
initiatives=[]; iid=1
for p in PILLARS:
    for (name,status,comp,sd,ed,budget,milestone,risk,impact) in INIT_DEFS[p]:
        initiatives.append({"id":f"INI-{iid:02d}","pillar":p,"name":name,"owner":OWNERS[p],
            "completion":comp,"status":status,"start":sd,"end":ed,"budget":budget,
            "milestone":milestone,"risk":risk,"impact":impact}); iid+=1
# the three closing within 90 days (2026-07-08 → 2026-10-06):
CLOSING=[i for i in initiatives if "2026-07-08"<i["end"]<="2026-10-06" and i["status"]!="مكتملة"]
assert len(CLOSING)==3, [i["end"] for i in initiatives]

AVG_COMPLETION=round(sum(i["completion"] for i in initiatives)/len(initiatives),1)

# ----------------------------------------------------------------------------
# 14  KPIs — 6 strategic + 8 operational, 12 monthly actuals each
# ----------------------------------------------------------------------------
COVERAGE=round(TOTAL_BEDS/TOTAL_DEMAND*100,1)                       # 43.1
RESIDENCY=round(TOTAL_OCC/TOTAL_DEMAND*100,1)                       # 39.5
OCC_RATE=round(TOTAL_OCC/TOTAL_BEDS*100,1)                          # 91.7
last12=inspections[-60:]
COMPLIANCE=round((1-sum(r["violations"] for r in last12)/sum(r["visits"] for r in last12))*100,1)
lic12=[r for r in licenses if r["month"]>= MONTHS24[12][2]]
DAYS_BLEND=round((sum(r["days_constr"] for r in lic12)+sum(r["days_ops"] for r in lic12))/(2*len(lic12)),1)
CLOSE_RATE=round(sum(r["req_closed"] for r in lic12)/sum(r["req_in"] for r in lic12)*100,1)
VISITS_PER_INSP=round(sum(r["visits"] for r in last12)/12/18,1)
FINES_MONTHLY_M=round(sum(r["fines"] for r in last12)/12/1_000_000,2)

def kpi_series(start,end,n=12,decimals=1,down=False):
    out=[]
    for i in range(n):
        v=start+(end-start)*(i/(n-1))+random.uniform(-abs(end-start)*0.06,abs(end-start)*0.06)
        out.append(round(v,decimals))
    out[-1]=round(end,decimals)
    return out

kpis=[
 dict(id="K1",name="نسبة تغطية الأسرّة المرخصة من إجمالي الطلب",level="استراتيجي",unit="٪",target=60.0,actual=COVERAGE,dir="أعلى أفضل",series=kpi_series(41.2,COVERAGE)),
 dict(id="K2",name="نسبة العمالة المقيمة في سكن مرخص",level="استراتيجي",unit="٪",target=45.0,actual=RESIDENCY,dir="أعلى أفضل",series=kpi_series(37.6,RESIDENCY)),
 dict(id="K3",name="معدل الامتثال في الجولات الرقابية",level="استراتيجي",unit="٪",target=88.0,actual=COMPLIANCE,dir="أعلى أفضل",series=kpi_series(80.8,COMPLIANCE)),
 dict(id="K4",name="متوسط مدة إصدار الرخصة",level="استراتيجي",unit="يوم",target=30.0,actual=DAYS_BLEND,dir="أقل أفضل",series=kpi_series(36.5,DAYS_BLEND)),
 dict(id="K5",name="نسبة إنجاز المبادرات الاستراتيجية",level="استراتيجي",unit="٪",target=70.0,actual=AVG_COMPLETION,dir="أعلى أفضل",series=kpi_series(48.0,AVG_COMPLETION)),
 dict(id="K6",name="رضا المستفيدين",level="استراتيجي",unit="٪",target=80.0,actual=81.0,dir="أعلى أفضل",series=kpi_series(76.0,81.0)),
 dict(id="K7",name="معدل الإشغال",level="تشغيلي",unit="٪",target=88.0,actual=OCC_RATE,dir="أقل أفضل",series=kpi_series(90.4,OCC_RATE)),
 dict(id="K8",name="نسبة إغلاق طلبات الترخيص",level="تشغيلي",unit="٪",target=85.0,actual=CLOSE_RATE,dir="أعلى أفضل",series=kpi_series(82.5,CLOSE_RATE)),
 dict(id="K9",name="عدد الجولات لكل مفتش شهرياً",level="تشغيلي",unit="جولة",target=90.0,actual=VISITS_PER_INSP,dir="أعلى أفضل",series=kpi_series(78.0,VISITS_PER_INSP)),
 dict(id="K10",name="نسبة المخالفات المصححة خلال 30 يوماً",level="تشغيلي",unit="٪",target=80.0,actual=76.0,dir="أعلى أفضل",series=kpi_series(69.5,76.0)),
 dict(id="K11",name="نسبة الرخص المصدرة إلكترونياً",level="تشغيلي",unit="٪",target=90.0,actual=93.0,dir="أعلى أفضل",series=kpi_series(85.5,93.0)),
 dict(id="K12",name="متوسط زمن الاستجابة للبلاغات",level="تشغيلي",unit="ساعة",target=6.0,actual=9.6,dir="أقل أفضل",series=kpi_series(12.4,9.6)),
 dict(id="K13",name="نسبة تجديد الرخص في الوقت المحدد",level="تشغيلي",unit="٪",target=85.0,actual=82.0,dir="أعلى أفضل",series=kpi_series(77.0,82.0)),
 dict(id="K14",name="متوسط الغرامات المحصلة شهرياً",level="تشغيلي",unit="مليون ريال",target=2.5,actual=FINES_MONTHLY_M,dir="أعلى أفضل",series=kpi_series(max(1.6,FINES_MONTHLY_M-0.7),FINES_MONTHLY_M,decimals=2)),
]

# ----------------------------------------------------------------------------
# 15  PREDICTIONS — 12-month forecast (3 scenarios) + 10 insight rows
# ----------------------------------------------------------------------------
SCEN={"متحفظ":{"d":9400,"s":1900},"أساسي":{"d":7200,"s":2800},"متفائل":{"d":5600,"s":4600}}
forecast={}
for scen,p in SCEN.items():
    rows=[]
    for i,(y,m,key,label) in enumerate(FCAST12):
        t=i+1
        d=round(TOTAL_DEMAND+p["d"]*t*(1+0.006*t))
        s=round(TOTAL_BEDS+p["s"]*t*(1+0.004*t))
        band=0.012+0.0016*t
        rows.append({"month":key,"month_ar":label,"demand":d,"supply":s,
            "d_lo":round(d*(1-band)),"d_hi":round(d*(1+band)),
            "s_lo":round(s*(1-band*0.7)),"s_hi":round(s*(1+band*0.7)),
            "gap":d-s})
    forecast[scen]=rows
BASE_GAP_END=forecast["أساسي"][-1]["gap"]

CLOSING90=sorted(CLOSING,key=lambda i:-i["completion"])[:3]
east_hot=[f for f in facilities if f["sector"]=="القطاع الشرقي" and f["occ_rate"]>95]
south_density=round(sum(h["density"] for h in hotspots if h["sector"]=="القطاع الجنوبي")/12)
south_viol_12=sum(r["violations"] for r in inspections[-60:] if r["sector"]=="القطاع الجنوبي")
lic_last3=[r for r in licenses if r["month"]>=MONTHS24[21][2]]
q_proj=round(sum(r["constr"]+r["ops"] for r in lic_last3)*1.18/10)*10

predictions=[
 dict(cat="طلب وعرض",conf=88,horizon="12 شهراً",priority="عالية",
   text=f"استمرار اتساع فجوة الإيواء لتبلغ نحو {BASE_GAP_END:,} سرير بنهاية يونيو 2027 وفق السيناريو الأساسي، مع نمو الطلب بوتيرة تفوق نمو الأسرّة المرخصة.",
   rec="تسريع طرح الفرص الاستثمارية للمجمعات الكبرى وتفعيل حوافز الترخيص للمطورين خلال الربعين القادمين."),
 dict(cat="طلب وعرض",conf=84,horizon="6 أشهر",priority="عالية",
   text=f"ضغط متوقع على الطاقة الاستيعابية في القطاع الشرقي؛ {len(east_hot)} منشأة تتجاوز نسبة إشغالها 95٪ ومتوسط إشغال القطاع 95.4٪.",
   rec="توجيه الرخص الجديدة نحو الشرقي وتخصيص مسار سريع لطلبات التوسعة القائمة."),
 dict(cat="رقابة",conf=86,horizon="3 أشهر",priority="عالية",
   text=f"القطاع الجنوبي يتطلب تكثيف الرقابة: متوسط كثافة النقاط الساخنة {south_density} نقطة و{south_viol_12:,} مخالفة خلال آخر 12 شهراً (الأعلى بين القطاعات).",
   rec="زيادة الجولات الميدانية في السلي والمنصورية 25٪ وتفعيل الحملة المشتركة مع الدفاع المدني."),
 dict(cat="مبادرات",conf=92,horizon="90 يوماً",priority="متوسطة",
   text="ثلاث مبادرات تقترب من الإغلاق خلال 90 يوماً: "+"، ".join(f"{i['name']} ({i['completion']}٪)" for i in CLOSING90)+".",
   rec="عقد مراجعات إغلاق أسبوعية للمبادرات الثلاث وتأمين الاعتمادات النهائية قبل سبتمبر 2026."),
 dict(cat="تراخيص",conf=81,horizon="ربع سنوي",priority="متوسطة",
   text=f"إسقاط الاتجاه الحالي (+18٪ سنوياً) يشير إلى نحو {q_proj:,} رخصة متوقعة في الربع القادم، ما قد يرفع زمن المعالجة إذا بقيت الطاقة التشغيلية ثابتة.",
   rec="رفع الطاقة الاستيعابية لفريق دراسة الطلبات 15٪ وتوسيع الإصدار الآلي للحالات منخفضة المخاطر."),
 dict(cat="رقابة",conf=77,horizon="6 أشهر",priority="متوسطة",
   text="نمط المخالفات يتحول من النظافة العامة نحو الاكتظاظ واشتراطات السلامة، وهما فئتان تمثلان معاً 55٪ من مخالفات آخر 12 شهراً.",
   rec="إعادة توجيه برامج التوعية نحو معايير الطاقة الاستيعابية وتجهيزات الحريق قبل موسم الصيف."),
 dict(cat="طلب وعرض",conf=74,horizon="12 شهراً",priority="متوسطة",
   text="تشير النماذج التنبؤية إلى نمو الطلب في القطاع الجنوبي بنحو 4٪ مدفوعاً بمشاريع البنية التحتية، مقابل تغطية حالية لا تتجاوز 34.8٪ وهي الأدنى بين القطاعات.",
   rec="تخصيص أراضٍ بلدية في محور الجنوب لمجمعات نموذجية ضمن المرحلة الثانية من مخطط المناطق النموذجية."),
 dict(cat="تراخيص",conf=79,horizon="6 أشهر",priority="منخفضة",
   text="استمرار تحسن زمن إصدار رخص التشغيل ليقترب من 20 يوماً بنهاية 2026 مع اكتمال منصة الترخيص الإلكتروني الموحد.",
   rec="اعتماد مؤشر زمن إصدار مستهدف جديد (18 يوماً) اعتباراً من الربع الأول 2027."),
 dict(cat="مبادرات",conf=71,horizon="6 أشهر",priority="منخفضة",
   text="مبادرتا الشراكة مع القطاع الخاص وملتقى الرياض متوقفتان منذ أكثر من ربعين، ما يعرّض مستهدف إنجاز الركيزة الاستثمارية لفجوة 12 نقطة.",
   rec="إعادة هيكلة نطاق المبادرتين أو إعادة جدولتهما رسمياً في محفظة 2027."),
 dict(cat="رقابة",conf=83,horizon="30 يوماً",priority="عالية",
   text="12 منشأة تنتهي رخصها قريباً وتظهر في الوقت ذاته درجات امتثال أقل من 75، وهي مرشحة للتحول إلى مواقع عالية الخطورة إذا لم تعالج.",
   rec="جدولة زيارات تحقق استباقية للمنشآت الاثنتي عشرة قبل انتهاء رخصها وربط التجديد بتصحيح الملاحظات."),
]

# ----------------------------------------------------------------------------
# GEO — sector polygons for the SVG map (0–100 space)
# ----------------------------------------------------------------------------
GEO={
 "القطاع الشمالي":{"poly":[[13,36],[17,18],[27,10],[40,10],[47,17],[49,28],[45,38],[38,46],[27,47],[18,44]],"label":[31,28]},
 "القطاع الشرقي":{"poly":[[52,26],[54,12],[63,5],[78,4],[91,9],[96,22],[92,36],[82,45],[68,48],[57,44],[51,36]],"label":[74,24]},
 "القطاع الأوسط":{"poly":[[41,49],[50,45],[57,50],[58,60],[51,66],[42,62],[39,55]],"label":[49,55]},
 "القطاع الغربي":{"poly":[[6,60],[13,52],[24,49],[36,52],[38,62],[34,73],[24,77],[12,73],[5,67]],"label":[21,63]},
 "القطاع الجنوبي":{"poly":[[44,67],[52,64],[57,55],[68,50],[80,54],[86,66],[81,82],[68,93],[53,94],[43,84],[40,74]],"label":[64,73]},
}

def _inside(x,y,poly):
    n=len(poly); j=n-1; c=False
    for i in range(n):
        xi,yi=poly[i]; xj,yj=poly[j]
        if ((yi>y)!=(yj>y)) and (x < (xj-xi)*(y-yi)/(yj-yi)+xi): c=not c
        j=i
    return c
def _clamp_into(pt_holder, key_x, key_y, sec_key):
    for o in pt_holder:
        poly=GEO[o[sec_key]]["poly"]
        cx=sum(p[0] for p in poly)/len(poly); cy=sum(p[1] for p in poly)/len(poly)
        x,y=o[key_x],o[key_y]; t=0.0
        while not _inside(x,y,poly) and t<1.0:
            t+=0.08
            x=o[key_x]+(cx-o[key_x])*t; y=o[key_y]+(cy-o[key_y])*t
        o[key_x]=round(x,1); o[key_y]=round(y,1)
_clamp_into(facilities,"x","y","sector")
_clamp_into(hotspots,"x","y","sector")

# ----------------------------------------------------------------------------
# ASSERTIONS — hard reconciliation checks
# ----------------------------------------------------------------------------
for nm,rows in [("occupation",[r[2] for r in demand_occupation]),
                ("econ",[r["total"] for r in demand_econ]),("sme",[r["total"] for r in demand_sme]),
                ("nat",[r["total"] for r in demand_nat]),("age",[r["total"] for r in demand_age]),
                ("sector",[r["total"] for r in demand_sector])]:
    assert sum(rows)==TOTAL_DEMAND, (nm,sum(rows))
for rows in [demand_econ,demand_sme,demand_nat,demand_age,demand_sector]:
    assert sum(r["blue"] for r in rows)==BLUE_TOTAL and sum(r["white"] for r in rows)==WHITE_TOTAL
assert sum(r[2] for r in demand_occupation if r[1]=="ياقات زرقاء")==BLUE_TOTAL
assert len(facilities)==140
assert sum(f["beds"] for f in facilities)==TOTAL_BEDS
assert sum(f["occupied"] for f in facilities)==TOTAL_OCC
for f in facilities:
    lo,hi=RANGES[f["type"]]; assert lo<=f["beds"]<=hi,(f["type"],f["beds"])
assert sum(c["count"] for c in viol_cats)==TOT_VIOL
assert sum(c["fines"] for c in viol_cats)==TOT_FINES
for c in viol_cats:
    assert sum(c["sme"].values())==c["count"] and sum(c["by_sector"].values())==c["count"]
assert sum(i["visits"] for i in inspectors)==TOT_VISITS
assert sum(i["violations"] for i in inspectors)==TOT_VIOL
y1=sum(r["constr"]+r["ops"] for r in licenses if r["month"]<MONTHS24[12][2])
y2=sum(r["constr"]+r["ops"] for r in licenses if r["month"]>=MONTHS24[12][2])
YOY=round((y2/y1-1)*100,1)
assert 14<=YOY<=22, YOY
assert len(hotspots)==40 and len(initiatives)==28 and len(kpis)==14 and len(predictions)==10
print("✓ all reconciliation assertions passed")
print(f"  demand={TOTAL_DEMAND:,} beds={TOTAL_BEDS:,} occupied={TOTAL_OCC:,} coverage={COVERAGE}% occupancy={OCC_RATE}%")
print(f"  visits24m={TOT_VISITS:,} violations24m={TOT_VIOL:,} fines24m={TOT_FINES:,} SAR closures={TOT_CLOSURES}")
print(f"  licenses YoY=+{YOY}% | compliance12m={COMPLIANCE}% | closeRate12m={CLOSE_RATE}% | initiatives avg={AVG_COMPLETION}%")
print(f"  base-scenario gap @ horizon = {BASE_GAP_END:,}")

# ----------------------------------------------------------------------------
# data.json for the dashboard
# ----------------------------------------------------------------------------
DATA={
 "meta":{"title":"لوحة معلومات السكن الجماعي للأفراد بمدينة الرياض","entity":"أمانة منطقة الرياض — وكالة التنمية الحضرية",
   "updated":"08 يوليو 2026 — 09:30 صباحاً","today":TODAY,
   "totals":{"demand":TOTAL_DEMAND,"beds":TOTAL_BEDS,"occupied":TOTAL_OCC,"gap":TOTAL_DEMAND-TOTAL_BEDS,
     "coverage":COVERAGE,"occupancy":OCC_RATE,"facilities":len(facilities),
     "blue":BLUE_TOTAL,"white":WHITE_TOTAL,"licensesYoY":YOY}},
 "sectors":SECTORS,
 "geo":GEO,
 "demand":{
   "occupation":[{"name":n,"collar":c,"count":v} for n,c,v in demand_occupation],
   "econ":demand_econ,"sme":demand_sme,"nat":demand_nat,"age":demand_age,"sector":demand_sector},
 "facilities":facilities,
 "licenses":licenses,
 "inspections":inspections,
 "violCats":viol_cats,
 "inspectors":inspectors,
 "hotspots":hotspots,
 "initiatives":initiatives,
 "pillars":PILLARS,
 "kpis":kpis,
 "kpiMonths":[lbl for _,_,_,lbl in month_seq(2025,7,12)],
 "forecast":forecast,
 "predictions":predictions,
 "facts":{
   "gapEnd":BASE_GAP_END,"eastHot":[{"name":f["name"],"occ":f["occ_rate"]} for f in sorted(east_hot,key=lambda z:-z["occ_rate"])[:6]],
   "eastHotCount":len(east_hot),"southDensity":south_density,"southViol12":south_viol_12,"qProj":q_proj,"bedsYoY":8.6,
   "closing90":[{"name":i["name"],"completion":i["completion"],"end":i["end"],"pillar":i["pillar"]} for i in CLOSING90]},
}
with open(os.path.join(HERE,"data.json"),"w",encoding="utf-8") as f:
    json.dump(DATA,f,ensure_ascii=False,separators=(",",":"))
print(f"✓ data.json written ({os.path.getsize(os.path.join(HERE,'data.json'))//1024} KB)")

# ----------------------------------------------------------------------------
# EXCEL WORKBOOK — labor_accommodation_data.xlsx
# ----------------------------------------------------------------------------
wb=Workbook(); wb.remove(wb.active)
HDR_FILL=PatternFill("solid",fgColor="0E5A43"); HDR_FONT=Font(name="Arial",bold=True,color="FFFFFF",size=11)

def sheet(name,headers,rows,widths=None):
    ws=wb.create_sheet(name); ws.sheet_view.rightToLeft=True
    ws.append(headers)
    for c in ws[1]:
        c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=Alignment(horizontal="center",vertical="center")
    for r in rows: ws.append(r)
    ws.freeze_panes="A2"
    for i,h in enumerate(headers,1):
        w=(widths[i-1] if widths and i<=len(widths) else max(14,min(46,len(str(h))+8)))
        ws.column_dimensions[get_column_letter(i)].width=w
    return ws

sheet("README",["الحقل","الوصف"],[
 ["الجهة","أمانة منطقة الرياض — وكالة التنمية الحضرية (بيانات تجريبية لأغراض العرض)"],
 ["إجمالي العمالة (الطلب)","1,420,000 عامل — يبلغ مجموع كل ورقة من أوراق الطلب هذا الرقم تماماً"],
 ["إجمالي الأسرّة المرخصة (العرض)","612,400 سرير — يساوي مجموع أسرّة ورقة Facilities تماماً"],
 ["فجوة الإيواء","807,600 سرير | نسبة التغطية 43.1٪"],
 ["معدل الإشغال","91.7٪ — الأسرّة المشغولة 561,571 موزعة على 140 منشأة"],
 ["النطاق الزمني","بيانات شهرية من يوليو 2024 إلى يونيو 2026، وتوقعات حتى يونيو 2027"],
 ["Demand_*","ست أوراق لتفصيل الطلب (المهن، القطاع الاقتصادي، حجم المنشأة، الجنسية، العمر، القطاع البلدي) ومجموع كل منها يساوي 1,420,000"],
 ["Facilities","140 منشأة سكن مرخصة بإحداثيات X/Y على خريطة القطاعات (فضاء 0–100)"],
 ["Licenses_Monthly","إصدار الرخص والطلبات لكل شهر × قطاع × نوع سكن (نمو سنوي ≈ +18٪)"],
 ["Inspections_Monthly","الجولات والمخالفات والغرامات والإغلاقات لكل شهر × قطاع"],
 ["Violation_Categories","فئات المخالفات وتوزيعها حسب حجم المنشأة والقطاع — مجموعها يطابق Inspections"],
 ["Inspectors / Hotspots","أداء 18 مفتشاً و40 نقطة ساخنة لكثافة المخالفات"],
 ["Initiatives / KPIs / Predictions","28 مبادرة على 4 ركائز استراتيجية، 14 مؤشر أداء، وتوقعات 12 شهراً بثلاثة سيناريوهات"],
 ["dim_*","أوراق مرجعية: القطاعات، الأحياء، أنواع السكن"],
 ["ملاحظة","الأرقام تجريبية متسقة داخلياً؛ لاستبدالها ببيانات حقيقية حدّث الأوراق مع الحفاظ على أسماء الأعمدة"],
],[28,110])

sheet("Demand_ByOccupation",["المجموعة المهنية","نوع الياقة","عدد العمالة"],
 [[n,c,v] for n,c,v in demand_occupation],[42,18,16])
sheet("Demand_ByEconomicSector",["القطاع الاقتصادي","عدد العمالة","ياقات زرقاء","ياقات بيضاء"],
 [[r["name"],r["total"],r["blue"],r["white"]] for r in demand_econ],[32,16,16,16])
sheet("Demand_BySMESize",["حجم المنشأة","عدد العمالة","ياقات زرقاء","ياقات بيضاء"],
 [[r["name"],r["total"],r["blue"],r["white"]] for r in demand_sme],[22,16,16,16])
sheet("Demand_ByNationality",["الجنسية","عدد العمالة","ياقات زرقاء","ياقات بيضاء"],
 [[r["name"],r["total"],r["blue"],r["white"]] for r in demand_nat],[20,16,16,16])
sheet("Demand_ByAge",["الفئة العمرية","عدد العمالة","ياقات زرقاء","ياقات بيضاء"],
 [[r["name"],r["total"],r["blue"],r["white"]] for r in demand_age],[18,16,16,16])
sheet("Demand_ByMunicipalSector",["القطاع البلدي","عدد العمالة","ياقات زرقاء","ياقات بيضاء"],
 [[r["name"],r["total"],r["blue"],r["white"]] for r in demand_sector],[22,16,16,16])
sheet("Facilities",["رقم المنشأة","اسم المنشأة","نوع السكن","القطاع البلدي","الحي","X","Y",
 "الأسرّة المرخصة","الأسرّة المشغولة","معدل الإشغال ٪","حالة الترخيص","تاريخ الترخيص","درجة الامتثال","مستوى الخطورة","آخر زيارة رقابية"],
 [[f["id"],f["name"],f["type"],f["sector"],f["district"],f["x"],f["y"],f["beds"],f["occupied"],
   f["occ_rate"],f["lic_status"],f["lic_date"],f["compliance"],f["risk"],f["last_visit"]] for f in facilities],
 [12,34,14,18,16,8,8,16,16,14,14,14,14,14,16])
sheet("Licenses_Monthly",["الشهر","القطاع البلدي","نوع السكن","رخص البناء الصادرة","رخص التشغيل الصادرة",
 "الطلبات المستلمة","الطلبات المغلقة","متوسط مدة إصدار رخصة البناء (يوم)","متوسط مدة إصدار رخصة التشغيل (يوم)"],
 [[r["month_ar"],r["sector"],r["type"],r["constr"],r["ops"],r["req_in"],r["req_closed"],r["days_constr"],r["days_ops"]] for r in licenses],
 [16,18,14,18,18,16,16,30,30])
sheet("Inspections_Monthly",["الشهر","القطاع البلدي","الزيارات الميدانية","المخالفات المحررة","الغرامات المحصلة (ريال)","المنشآت المغلقة"],
 [[r["month_ar"],r["sector"],r["visits"],r["violations"],r["fines"],r["closures"]] for r in inspections],
 [16,18,18,18,22,16])
sheet("Violation_Categories",["فئة المخالفة","عدد المخالفات","إجمالي الغرامات (ريال)",
 "متناهية الصغر","صغيرة","متوسطة","كبيرة"]+SECTORS,
 [[c["name"],c["count"],c["fines"]]+[c["sme"][s] for s in SME_ORDER]+[c["by_sector"][s] for s in SECTORS] for c in viol_cats],
 [38,16,20,14,12,12,12,16,16,16,16,16])
sheet("Inspectors",["المفتش","القطاع البلدي","عدد الجولات","المخالفات المحررة","المسافة المقطوعة (كم)"],
 [[i["name"],i["sector"],i["visits"],i["violations"],i["distance"]] for i in inspectors],[24,18,14,18,20])
sheet("Hotspots",["X","Y","القطاع البلدي","كثافة المخالفات"],
 [[h["x"],h["y"],h["sector"],h["density"]] for h in hotspots],[10,10,20,18])
sheet("Initiatives",["رقم المبادرة","الركيزة","اسم المبادرة","الجهة المسؤولة","نسبة الإنجاز ٪","الحالة",
 "تاريخ البداية","تاريخ النهاية المستهدف","الميزانية (ريال)","المعلم القادم","مستوى المخاطر","الأثر المتوقع"],
 [[i["id"],i["pillar"],i["name"],i["owner"],i["completion"],i["status"],i["start"],i["end"],i["budget"],i["milestone"],i["risk"],i["impact"]] for i in initiatives],
 [12,22,46,28,14,12,14,20,16,42,14,44])
sheet("KPIs",["المؤشر","النوع","الوحدة","المستهدف","المتحقق","الاتجاه"]+[lbl for _,_,_,lbl in month_seq(2025,7,12)],
 [[k["name"],k["level"],k["unit"],k["target"],k["actual"],k["dir"]]+k["series"] for k in kpis],
 [42,12,12,12,12,12]+[13]*12)
pred_rows=[[p["text"],p["cat"],p["conf"],p["horizon"],p["rec"],p["priority"]] for p in predictions]
sheet("Predictions",["الرؤية/التنبؤ","الفئة","مستوى الثقة ٪","الأفق الزمني","التوصية","الأولوية"],pred_rows,[80,14,14,14,70,12])
fc_rows=[]
for scen,rows in forecast.items():
    for r in rows:
        fc_rows.append([scen,r["month_ar"],r["demand"],r["d_lo"],r["d_hi"],r["supply"],r["s_lo"],r["s_hi"],r["gap"]])
sheet("Forecast_Series",["السيناريو","الشهر","الطلب المتوقع","الحد الأدنى (طلب)","الحد الأعلى (طلب)",
 "العرض المتوقع","الحد الأدنى (عرض)","الحد الأعلى (عرض)","الفجوة المتوقعة"],fc_rows,[12,16,16,18,18,16,18,18,16])
sheet("dim_sector",["sector_id","القطاع البلدي","sector_en","centroid_x","centroid_y","color_code"],
 [[SEC_EN[s],s,SEC_EN[s].title(),GEO[s]["label"][0],GEO[s]["label"][1],"#0E5A43"] for s in SECTORS],[12,20,12,12,12,12])
dist_rows=[]
for s in SECTORS:
    for d in DISTRICTS[s]:
        n=len([f for f in facilities if f["district"]==d and f["sector"]==s])
        dist_rows.append([f"D-{SEC_EN[s][:2]}-{DISTRICTS[s].index(d)+1}",d,s,random.choice(["مرتفعة","متوسطة","منخفضة"]),
                          "مرتفع" if s=="القطاع الجنوبي" else random.choice(["متوسط","منخفض"]),n])
sheet("dim_district",["district_id","الحي","القطاع البلدي","كثافة العمالة التقديرية","مستوى الخطورة","عدد المنشآت"],dist_rows,[14,18,20,22,16,14])
sheet("dim_accommodation_type",["accommodation_type_id","نوع السكن","الوصف"],
 [["COMPOUND","مجمع سكني","مجمعات كبرى مخصصة لإسكان العمالة بطاقة 3,000–22,000 سرير"],
  ["BUILDING","مبنى سكني","مبانٍ قائمة مرخصة للسكن الجماعي بطاقة 200–1,200 سرير"],
  ["CABINS","كبائن متنقلة","كبائن مؤقتة داخل مواقع المشاريع بطاقة 80–600 سرير"]],[22,16,60])

out=os.path.join(HERE,"labor_accommodation_data.xlsx")
wb.save(out)
print(f"✓ workbook written: {out} ({os.path.getsize(out)//1024} KB, {len(wb.sheetnames)} sheets)")
print("  sheets:", ", ".join(wb.sheetnames))
